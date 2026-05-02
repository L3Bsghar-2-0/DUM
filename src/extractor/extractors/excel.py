from __future__ import annotations
import re
from pathlib import Path
from datetime import datetime, date, time
import openpyxl
from rapidfuzz import process as fuzz_process
from extractors.base import ExtractionResult

ROW_FIELD_MAP: dict[int, str] = {
    12: 'gaz_volume_nm3',
    14: 'gaz_debit_nm3h',
    15: 'elec_auxiliaire_kwh',
    16: 'puissance_brute_kw',
    17: 'heures_fonctionnement',
    18: 'energie_alternateur_kwh',
    19: 'energie_reactive_kvarh',
    20: 'vitesse_rpm',
    21: 'facteur_puissance',
    22: 'voltage_v',
    23: 'courant_phase1_a',
    24: 'courant_phase2_a',
    25: 'courant_phase3_a',
    26: 'eg_debit_m3h',
    27: 'eg_temp_entree_c',
    28: 'eg_temp_sortie_c',
    29: 'eg_energie_kwh',
    30: 'eg_puissance_kw',
    31: 'ec_recup_debit_m3h',
    32: 'ec_recup_temp_entree_c',
    33: 'ec_recup_temp_sortie_c',
    34: 'ec_recup_energie_kwh',
    35: 'ec_recup_puissance_kw',
    36: 'ec_alpha_sani_debit_m3h',
    37: 'ec_alpha_sani_temp_entree_c',
    38: 'ec_alpha_sani_temp_sortie_c',
    39: 'ec_alpha_sani_energie_kwh',
    40: 'ec_alpha_sani_puissance_kw',
    43: 'ec_alpha_debit_m3h',
    44: 'ec_alpha_temp_entree_c',
    45: 'ec_alpha_temp_sortie_c',
    46: 'ec_alpha_energie_kwh',
    47: 'ec_alpha_puissance_kw',
    50: 'ec_gamma_debit_m3h',
    51: 'ec_gamma_temp_entree_c',
    52: 'ec_gamma_temp_sortie_c',
    53: 'ec_gamma_energie_kwh',
    54: 'ec_gamma_puissance_kw',
    57: 'rendement_electrique_pct',
    58: 'rendement_thermique_pct',
    59: 'rendement_total_pct',
    60: 'steg_achat_kwh',
    61: 'steg_vente_kwh',
    62: 'production_positive_kwh',
    63: 'production_negative_kwh',
}

FRENCH_LABELS: dict[str, str] = {
    "consommation du gaz naturel moteur en nm3": "gaz_volume_nm3",
    "debit du gaz naturel moteur en nm3/h": "gaz_debit_nm3h",
    "energie electrique en kwh": "elec_auxiliaire_kwh",
    "puissance electrique brute en kw": "puissance_brute_kw",
    "heure de fonctionnement": "heures_fonctionnement",
    "energie electrique au borne de l'alternateur en kwh": "energie_alternateur_kwh",
    "energie reactive en kvarh": "energie_reactive_kvarh",
    "vitesse en rpm": "vitesse_rpm",
    "facteur de puissance": "facteur_puissance",
    "voltage en v": "voltage_v",
    "courant: phase 1 en a": "courant_phase1_a",
    "courant: phase 2 en a": "courant_phase2_a",
    "courant: phase 3 en a": "courant_phase3_a",
    "rendement electrique %": "rendement_electrique_pct",
    "rendement thermique %": "rendement_thermique_pct",
    "rendement total %": "rendement_total_pct",
    "energie positive steg kwh": "steg_achat_kwh",
    "energie negative steg kwh": "steg_vente_kwh",
    "energie positive production kwh": "production_positive_kwh",
    "energie negative production kwh": "production_negative_kwh",
}

_PCI_RE = re.compile(r'pci\s*\(thermie/nm3\)[:\s]*([0-9.,]+)', re.IGNORECASE)
_COSPHI_RE = re.compile(r'cos.*?=\s*([0-9.,]+)', re.IGNORECASE)


def _parse_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '.').strip()
    try:
        return float(s)
    except ValueError:
        return None


def _combine_datetime(d, t) -> datetime | None:
    if isinstance(d, datetime):
        dt_date = d.date()
    elif isinstance(d, date):
        dt_date = d
    else:
        return None
    if isinstance(t, time):
        dt_time = t
    elif isinstance(t, datetime):
        dt_time = t.time()
    else:
        dt_time = time(0, 0)
    return datetime.combine(dt_date, dt_time)


class ExcelExtractor:
    def extract(self, path: Path) -> list[ExtractionResult]:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        sheet_name = next(
            (s for s in wb.sheetnames if 'bilan' in s.lower() or 'total' in s.lower()),
            wb.sheetnames[0]
        )
        ws = wb[sheet_name]

        pci = 9.082
        cos_phi = 1.0
        warnings: list[str] = []

        for row_idx in range(1, 9):
            cell_val = ws.cell(row_idx, 2).value
            if cell_val is None:
                continue
            s = str(cell_val)
            m = _PCI_RE.search(s)
            if m:
                try:
                    pci = float(m.group(1).replace(',', '.'))
                except ValueError:
                    pass
            m2 = _COSPHI_RE.search(s)
            if m2:
                try:
                    cos_phi = float(m2.group(1).replace(',', '.'))
                except ValueError:
                    pass

        verified_map: dict[int, str] = dict(ROW_FIELD_MAP)
        label_keys = list(FRENCH_LABELS.keys())
        for row_idx, field in ROW_FIELD_MAP.items():
            label_cell = ws.cell(row_idx, 2).value
            if label_cell is None:
                continue
            normalized = str(label_cell).lower().strip()
            match = fuzz_process.extractOne(normalized, label_keys, score_cutoff=70)
            if match:
                fuzzy_field = FRENCH_LABELS[match[0]]
                if fuzzy_field != field:
                    warnings.append(
                        f"Row {row_idx}: fuzzy label '{match[0]}' maps to '{fuzzy_field}', "
                        f"using hardcoded '{field}'"
                    )

        max_col = ws.max_column or 1
        results: list[ExtractionResult] = []

        for col_idx in range(5, max_col + 1):
            date_val = ws.cell(10, col_idx).value
            time_val = ws.cell(11, col_idx).value
            if date_val is None:
                continue
            ts = _combine_datetime(date_val, time_val)

            record_data: dict = {}
            for row_idx, field in verified_map.items():
                cell = ws.cell(row_idx, col_idx)
                val = _parse_float(cell.value)
                if val is not None:
                    record_data[field] = val

            if not record_data:
                continue

            results.append(ExtractionResult(
                timestamp=ts,
                source_file=path.name,
                source_type="excel",
                pci_thermie_nm3=pci,
                cos_phi=cos_phi,
                extraction_warnings=list(warnings),
                confidence_score=len(record_data) / len(verified_map),
                **record_data,
            ))

        wb.close()
        return results
